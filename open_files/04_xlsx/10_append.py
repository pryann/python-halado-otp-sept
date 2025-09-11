import openpyxl
import os


def get_max_id(file):
    max_id = 0

    try:
        if os.path.exists(file):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Fejléc olvasása
            columns = [cell.value for cell in sheet[1]]
            if "id" in columns:
                id_index = columns.index("id")
                for row in sheet.iter_rows(
                    min_row=2, max_col=id_index + 1, values_only=True
                ):
                    if row[id_index] and isinstance(row[id_index], int):
                        max_id = max(max_id, row[id_index])

    except Exception as e:
        print(f"Error reading file for max ID: {e}")

    return max_id


def add_id_field(file, rows):
    max_id = get_max_id(file)

    for row in rows:
        max_id += 1
        row["id"] = max_id

    fields = list(rows[0].keys())

    return fields, rows


def append_xlsx(file, rows):
    file_exists = os.path.isfile(file)
    fields, rows = add_id_field(file, rows)

    if file_exists:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(fields)

    for row in rows:
        sheet.append(list(row.values()))

    workbook.save(file)


if __name__ == "__main__":
    new_rows = [
        {
            "first_name": "Robert",
            "last_name": "Smith",
            "email_address": "robert.smith@company.com",
            "gender": "Male",
            "yearly_salary": 110000,
            "years_of_experience": 7,
        },
        {
            "first_name": "Maria",
            "last_name": "Garcia",
            "email_address": "maria.garcia@company.com",
            "gender": "Female",
            "yearly_salary": 95000,
            "years_of_experience": 5,
        },
    ]

    filename = "./files/employees-with-header.xlsx"
    append_xlsx(filename, new_rows)
    print(f"Appended {len(new_rows)} rows to {filename}")
